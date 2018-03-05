# -*- coding: utf-8 -*-
"""
File Name: init_femb.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 7/15/2016 11:47:39 AM
Last modified: 10/18/2016 4:37:37 PM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
import numpy as np
import struct
import os
from scripts.detect_peaks import detect_peaks
import matplotlib.pyplot as plt
from scipy import stats
import sys
import glob
import pickle
import re
from scripts.Data_Analysis import Data_Analysis
import time
import glob
import pickle
import warnings
from user_settings import user_editable_settings
settings = user_editable_settings()

analyze = Data_Analysis()

####################################################################################################################################################
#This function should be run after the RMS file with the slope data has already been made.  It takes in the root folder, the specific directory
#to look through, the existing workbook to save to, and the total number of chips.  If will read the binary file created for RMS data analysis.
#It will find the mean and RMS of the data, and convert it to other units, and format it all for final analysis in the RMS spreadsheet.

def rms_process(test_dir, wb):
    rms_data_file = test_dir + "\\pedestal.dat"

#Name the sheet based on the directory you're looking into for the above file
    config_file = (glob.glob(test_dir + "*chip_settings*")[0])
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
        baseline = (data["base"])
        gain = (data["gain"])
        peak = (data["peak"])
        
    sheet_title = gain[0:4] + "," + peak[0:3] + "," + baseline[0:3]
    print (sheet_title)

    with open(rms_data_file, 'rb') as f:
        raw_data = f.read()
        
    start_of_packet = (b"\xde\xad\xbe\xef")
        
    start_of_chip = []
    for m in re.finditer(start_of_packet, raw_data):
        start_of_chip.append(m.start())
        
    chip_num = settings.chip_num
    if (len(start_of_chip) != chip_num):
        print ("RMS Analysis--> {} doesn't have {} chips in the file!".format(rms_data_file, chip_num))
        
    separated_by_chip = [(raw_data[start_of_chip[0] : start_of_chip[1]]),
                         (raw_data[start_of_chip[1] : start_of_chip[2]]),
                         (raw_data[start_of_chip[2] : start_of_chip[3]]),
                         (raw_data[start_of_chip[3] : ])]
        
    
    chip_data = [[],[],[],[]]
    for i in range(chip_num):
        chip_data[i] = analyze.UnpackData(path = "bytes", data = separated_by_chip[i], return_data = True)


#Histograms are plotted in the folder for each configuration
    plot_dir = test_dir + "\\Histograms\\"

    try: 
        os.makedirs(plot_dir)
    except OSError:
        if os.path.exists(plot_dir):
            pass
#Mean, Standard Deviation and stuck bit arrays that get passed to the final Excel spreadsheet
    std_np = []
    mean_np = []
    stuck_np = []

#Positioning for histogram text
    av_text = (.8,-.14)
    mode_text = (.8,-.19)
    std_text = (-.10,-.24)
    samples_text = (-.10,-.14)
    samples_within_text = (-.10,-.19)
    st_text = (.8,-.24)
    stuck_text = (.4,-.24)

    for chip in range(chip_num):
        for chn in range(16):

#For a given chip and channel's data, read and find the following parameters
            np_data = np.array(chip_data[chip][chn])
            datamean = np.mean(np_data)
            mean_np.append (datamean)
            std = np.std(np_data)
            std_np.append (std)
            mode = stats.mstats.mode(np_data)
            total_num = len(np_data)

            low_range = datamean - (3*std)
            high_range = datamean + (3*std)

            num_within = 0

#As part of stuck bit check, finds the amount of samples within 3 sigma

            for i in range(total_num):
                if (np_data[i] > low_range) and (np_data[i] < high_range):
                    num_within += 1
             
            maximum=max(np_data)
            minimum=min(np_data)

            plot_path = plot_dir + "Chip" + str(chip+1) + "_Ch" + str(chn)

            bins = maximum - minimum
            if (bins < 1):
                bins = 1
            

            plt.figure(figsize=(12,8))
            ax = plt.subplot(1,1,1)  
            ax.hist(np_data,bins=bins)
            ax.set_xlabel("ADC Counts")
            ax.set_ylabel("Occurences")
            ax.set_title("ADC Count Distribution for Chip "+str(chip+1)+", Channel "+str(chn))

# Shrink current axis's height by 10% on the bottom
            box = ax.get_position()
            ax.set_position([box.x0, box.y0 + box.height * 0.1,
            box.width, box.height * 0.9])

#Self explanatory, it takes the statistics found above and plots them in the designated place.  'axes fraction' lets you give the coordinates
#in terms of the axes, where 0 is one extreme and 1 is the other.  Without that, the coordinates would be in data format, which is harder to use

            plt.annotate("Average = "+str(round(datamean,2)), xy=av_text,  xycoords='axes fraction',
            xytext=av_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            plt.annotate("Mode = "+str(int(mode[0][0])), xy=mode_text,  xycoords='axes fraction',
            xytext=mode_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            plt.annotate("Standard deviation = "+str(round(std,2)), xy=std_text,  xycoords='axes fraction',
            xytext=std_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            plt.annotate("Total samples = "+str(total_num), xy=samples_text,  xycoords='axes fraction',
            xytext=samples_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            plt.annotate("Total samples in 3 sigma = "+str(num_within), xy=samples_within_text,  xycoords='axes fraction',
            xytext=samples_within_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            plt.annotate("S/T Ratio = "+str(round(float(num_within)/float(total_num),6)), xy=st_text,  xycoords='axes fraction',
            xytext=st_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            r64 =  int(mode[0][0])%64
            if (r64>15) and (r64<50) and ((float(num_within)/float(total_num))>0.995):
                stuck_np.append(1)
            else: 
                stuck_np.append(0)
                plt.annotate("Flagged as Stuck", xy=stuck_text,  xycoords='axes fraction',
                xytext=stuck_text, textcoords='axes fraction',
                horizontalalignment='left', verticalalignment='left', color='red',
                )

            plt.savefig (plot_path+".jpg")
            plt.close()

#Does the stuck bit analysis, finding out whether the mean/mode is close to a multiple of 64, which is indicative of a stuck bit
#A 1 means it looks good, a 0 means it could be a stuck bit

    std_np = np.resize(std_np,[chip_num,16])
    mean_np = np.resize(mean_np,[chip_num,16]) 
    stuck_np =  np.resize(stuck_np,[chip_num,16])
    
#Search for the correct sheet in the workbook passed to this function
    ws = wb.get_sheet_by_name(name = sheet_title)

    center = Alignment(horizontal='center')

    title = ("Baseline = {}, Gain = {}, Peaking Time = {}".format(baseline, gain, peak))
#First two rows are for title blocks
    ws.merge_cells('B1:Q1')
    ws['B1'].value = title
    ws['B1'].alignment = center

#Make the formatting for each new chunk of data (The Chip# and Channel# cells as well as the merged title.  Send the number of the chunk (0, 1, 2),
#the number of chips being analyzed, the title you want, and ws.  These must match up with the actual data bring put in these chunks below
    make_title(0,4,"Mean Value (ADC Counts)",ws)
    make_title(1,4,"RMS Width (ADC Counts)",ws)
    make_title(2,4,"Stuck Bit Matrix - A 1 indicates that a stuck bit was detected for that channel",ws)
    make_title(5,4,"Mean value in mV",ws)
    make_title(6,4,"RMS in electrons",ws)
    
#Grab the existing slope data for each channel as found by the previous script and put it in a matrix to use
    mv_slope = []
    electron_slope = []

    for chip_id in range(chip_num):
        for chn in range(16):
            mv_slope.append(ws['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(3*(chip_num+3)))].value)
        
            electron_slope.append(ws['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(4*(chip_num+3)))].value)

    mv_slope = np.resize(mv_slope, [chip_num,16])
    electron_slope = np.resize(electron_slope, [chip_num,16])


#    Couldn't get borders to work, maybe at a later time
#    ws.cell('B2').border = Border(outline=Side(border_style='thick'))]


#Formatting of data

    for chip_id in range(chip_num):
        for chn in range(16):
            if ((mv_slope[chip_id,chn] != None) and (electron_slope[chip_id,chn])):
    #First chunk is Mean value in ADC counts
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4)].value=mean_np[chip_id,chn]
    #Second chunk is Standard Deviation in ADC counts
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chip_num+3))].value=std_np[chip_id,chn]
    #Third chunk is the binary stuck bit array
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(2*(chip_num+3)))].value=stuck_np[chip_id,chn]
    #Sixth chunk is Mean value in mV (baseline)
                ws['{}{}'.format(chr(ord('A') + chn + 1),
                                 chip_id+4+(5*(chip_num+3)))].value=mean_np[chip_id,chn]/mv_slope[chip_id,chn]
    #Seventh chunk is Standard Deviation in electrons (ENC)
                ws['{}{}'.format(chr(ord('A') + chn + 1),
                                 chip_id+4+(6*(chip_num+3)))].value=std_np[chip_id,chn]/electron_slope[chip_id,chn]

#Choose which chunks of data you want to mark as stuck or not.  You need to pass the stuck matrix, total chips, which chunk to analyze and ws
    for chunk in range(3):
        mark_stuck_bits(stuck_np,chip_num,chunk,ws)

    for chunk in range(6,7):
        mark_stuck_bits(stuck_np,chip_num,chunk,ws)

####################################################################################################################################################
#This it the first function to run, it takes in the root folder, the directory to look through, whether the internal or FPGA DAC was used, the
#workbook and the number of chips.  It looks through the data that was pulsed for each DAC value and records the peak value in the Pulse
#spreadsheet.

def gain_process(directory, intdac, wb):
    ws = wb.active
    ws = wb.create_sheet(0)

#Specify whether it's the FPGA or internal calibration so it knows which directory to look into and how many steps there were
    if intdac == 1:
        msb = 64
        calidir ="cali_intdac\\"
        init = "intdac_"   
        print ("Internal DAC Pulse")
    else:
        msb = 32
        calidir ="cali_fpgadac\\"
        init = "fpgadac_"
        print ("External FPGA DAC Pulse")
        
    config_file = (glob.glob(directory + "*chip_settings*")[0])
    
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
        baseline = (data["base"])
        gain = (data["gain"])
        peak = (data["peak"])

#Build the title of the sheet
    sheet_title = gain[0:4] + "," + peak[0:3]
    print (sheet_title)
    ws.title = sheet_title

    for dacvalue in range(msb):

#Find the total length of raw data file in bytes and read it into memory
        raw_data_file = directory + calidir + init + "%x"%dacvalue + ".dat"
        with open(raw_data_file, 'rb') as f:
            raw_data = f.read()
            
        start_of_packet = (b"\xde\xad\xbe\xef")
            
        start_of_chip = []
        for m in re.finditer(start_of_packet, raw_data):
            start_of_chip.append(m.start())
            
        chip_num = settings.chip_num    
        if (len(start_of_chip) != chip_num):
            print ("PULSE Analysis--> {} doesn't have {} chips in the file!".format(raw_data_file, chip_num))
            
        separated_by_chip = [(raw_data[start_of_chip[0] : start_of_chip[1]]),
                             (raw_data[start_of_chip[1] : start_of_chip[2]]),
                             (raw_data[start_of_chip[2] : start_of_chip[3]]),
                             (raw_data[start_of_chip[3] : ])]
            
        
        chip_data = [[],[],[],[]]
        for i in range(chip_num):
            chip_data[i] = analyze.UnpackData(path = "bytes", data = separated_by_chip[i], return_data = True)

#Formatting of titles and labels
        center = Alignment(horizontal='center')

        title = ("Baseline = {}, Gain = {}, Peaking Time = {}".format(baseline, gain, peak))
#Set up the title blocks
        ws.merge_cells('B1:Q1')
        ws['B1'].value = title
        ws['B1'].alignment = center

#Current way of doing the make_title function, but with a title that changes with each DAC step
        row1=2+(dacvalue*(chip_num+3))
        ws.merge_cells(start_row=row1,start_column=2,end_row=row1,end_column=17)
        ws['{}{}'.format("A",row1)].value = str(dacvalue)
        ws['{}{}'.format("B",row1)].value = "Average Peak Height (ADC Counts) for DAC step " + str(dacvalue+1)
        ws['{}{}'.format("B",row1)].alignment = center


        for chip in range(chip_num):
            for chn in range(16):

#Pull out one channel's array, get 
                np_data = np.array(chip_data[chip][chn])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", category=RuntimeWarning)
                    pedmean = np.nanmean(np_data[0:2048])
                maxmean = np.max(np_data[0:2048])

#Find peaks while inputting the "minimum peak height", which must be higher than half the max value found 
#and "minimum peak distance".  Returns the index of the peaks
                peaks_index = detect_peaks(x=np_data, mph=abs((maxmean+pedmean)/2), mpd=(500/2)) 
                peaks_value = []

#Go from index to the actual value
                for i in peaks_index :
                    peaks_value.append(np_data[i])

#If there were any peaks, average them
                if len(peaks_value) != 0 :
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", category=RuntimeWarning)
                        peaksmean = np.nanmean(peaks_value[0:-1])

#Or else just return the mean value of the baseline                    
                else:
                    peaksmean = pedmean
                chip_data[chip][chn] = peaksmean

        ft = Font(bold=True)
        #write to excel file
        for chip_id in range(chip_num):
            for chn in range(16):
                if (chip_id == 0):

#Quick way of making the Channel cells
                    vl = "Channel " + str(chn) 
                    ws['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].value = vl
                    ws['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].font = ft

#Writing the actual data
                ws['{}{}'.format(chr(ord('A') + chn + 1), 
                   chip_id+4+(dacvalue*(chip_num+3)))].value = chip_data[chip_id][chn]

#Quick way of making the Chip cells
            ws['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].font = ft

        print ("DAC Step: %x, Average Amplitude for Chip 1, Channel 0: %d"%(dacvalue, chip_data[1][0]))

#Make the formatting for each new chunk of data (The Chip# and Channel# cells as well as the merged title.  Takes in the number of the chunk (0, 1, 2),
#the number of chips being analyzed, the title you want, and ws.
def make_title(num,chip_num,title,ws):
        center = Alignment(horizontal='center')
        ft = Font(bold=True)
        title_row=2+(num*(chip_num+3))
        ws.merge_cells(start_row=title_row,start_column=2,end_row=title_row,end_column=17)
        ws['{}{}'.format("B", title_row)].value = title
        ws['{}{}'.format("B", title_row)].alignment = center

        for chip_id in range(chip_num):
#Adds the "Chip #" Field
            ws['{}{}'.format("A", 4+chip_id+(num*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws['{}{}'.format("A", 4+chip_id+(num*(chip_num+3)))].font = ft

        for chn in range(16):
#Adds the "Channel #" field            
            ws['{}{}'.format(chr(ord('A') + chn + 1),3+(num*(chip_num+3)))].value = "Channel " + str(chn)
            ws['{}{}'.format(chr(ord('A') + chn + 1),3+(num*(chip_num+3)))].font = ft

#Marks each data point as stuck or not.  You need to pass the stuck matrix, total chips, which chunk to analyze and ws
def mark_stuck_bits(stuck_np,chip_num,chunk,ws):

    redFill = PatternFill(start_color='FFC7CE', end_color='CEC7FF', fill_type='solid')
    redFont = Font(color='9C0006')

    for chip_id in range(chip_num):
        for chn in range(16):

#Makes the cells red if they represent a channel with stuck bits
            stuck = stuck_np[chip_id,chn]
            if (stuck == 0):
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chunk*(chip_num+3)))].fill=redFill
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chunk*(chip_num+3)))].font=redFont

#This is the final function to be run.  It takes the existing and completed RMS workbook, the number of chips, and the root folder.  It uses RMS
#data to go through and plot ENC vs. peaking time for each channel, and put all those plots in a folder called "ENC Plots" in the root folder.
def ENC_plots(wb,chip_num,plot_dir):
    print ("Plotting ENC data")
    ws = wb.active
    title = ws['B1'].value

#Remove the gain and peaking time, since it's variable
    subtitle = title[0:26]+title[65:150] 
    gains = ['04.7', '07.8', '14.0', '25.0']
    peaking_times = ['0.5', '1.0', '2.0', '3.0']
    plot_peak = [0.5,1,2,3]
    baselines = ['200', '900']

    subtitle_text = (.5,.94)    
    stuck_text = (.5,-.22)

    try: 
        os.makedirs(plot_dir)
    except OSError:
        if os.path.exists(plot_dir):
            pass

    for chip_id in range(chip_num):
        print ("Chip"+str(chip_id))
        for chn in range(16):
            for index,base in enumerate(baselines):

#Each baseline gets its own plot, since it gets messy if you try to have too many on one
                plt.figure(figsize=(12,8))
                ax = plt.subplot(1,1,1)                          
                ax.set_xlabel("Peaking Time (us)")
                ax.set_ylabel("ENC (electrons)")
                ax.set_title("ENC vs. Peaking Time for Chip "+str(chip_id)+", Channel "+str(chn)+", Baseline = "+str(base)+" mV")
                plot_path = plot_dir+"Chip"+str(chip_id)+"_Channel"+str(chn)+"_Baseline"+str(base)
                for index,gain in enumerate(gains):
                    ENC = []
                    for index,peaking_time in enumerate(peaking_times):

#Build the sheet name from the baseline, gain and peaking time parameter and find it
                        sheet = gain+","+peaking_time+","+base
                        ws = wb.get_sheet_by_name(name = sheet)

#On each sheet, the 6th chunk will have the ENC in electrons, so that channel's value is grabbed, as well as the fill color
                        ENC.append(ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(6*(chip_num+3)))].value)
                        font_color = ws['{}{}'.format(chr(ord('A') + chn + 1),
                                        chip_id+4+(6*(chip_num+3)))].fill.start_color.index

#If the fill color is charactaristic of a cell that has a stuck bit, plot it with an X.  If not, just a normal scatter plot
                        if (font_color == "00FFC7CE"):
                            ax.scatter(plot_peak[index], ENC[index], marker = 'x', s = 25, color = 'r')
                        else: 
                            ax.scatter(plot_peak[index], ENC[index], marker = '.', s =2, color = 'k')

#Once done, connect the dots
                    ax.plot(plot_peak, ENC, label=str(gain)+" mV/fC Gain")

                # Shrink current axis's height by 10% on the bottom
                box = ax.get_position()
                ax.set_position([box.x0, box.y0 + box.height * 0.1,
                box.width, box.height * 0.9])

#Format the auxilliary information around the plot
                plt.annotate(subtitle[0:50]+'\n'+subtitle[50:150], xy=subtitle_text,  xycoords='axes fraction',
                xytext=subtitle_text, textcoords='axes fraction',
                horizontalalignment='center', verticalalignment='center',
                )
                plt.annotate("Red X means there was a stuck bit detected during the RMS analysis", xy=stuck_text,  xycoords='axes fraction',
                xytext=stuck_text, textcoords='axes fraction',
                horizontalalignment='center', verticalalignment='center', color = 'r',
                )
                
                ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1),ncol=4, fontsize=11)
                plt.savefig (plot_path+".jpg")
                plt.close()
