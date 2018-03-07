# -*- coding: utf-8 -*-
"""
Created on Fri Jun  2 11:50:17 2017

@author: vlsilab2
"""

import os
import sys
import struct
import matplotlib.pyplot as plt
plt.switch_backend('agg')
import pickle
import numpy as np
from femb_python.test_measurements.quadFeTestCold.user_settings import user_editable_settings
from openpyxl import Workbook, load_workbook
import warnings
from femb_python.test_measurements.quadFeTestCold.scripts.int_dac_fit import int_dac_fit
from femb_python.test_measurements.quadFeTestCold.scripts.linear_fit_m import linear_fit
import matplotlib.patches as mpatches
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from femb_python.test_measurements.quadFeTestCold.scripts.detect_peaks import detect_peaks
from femb_python.test_measurements.quadFeTestCold.scripts.plotting import plot_functions
from scipy import stats
import math
from femb_python.test_measurements.quadFeTestCold.scripts.Data_Analysis3 import Data_Analysis3
settings = user_editable_settings()
import warnings
with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=RuntimeWarning)

class Data_Analysis2:
    
    #Puts the channel designations where they should be, as long as you know how many discrete sections there are and what the spacing between them is
    def setup_spreadsheet(self, ws, sections, rows, start):
        for i in range(sections):
            ws.cell(row = (rows * i) + start, column = 1).value = "Channel"
            for j in range(16):
                ws.cell(row = j + (rows * i) + start + 1, column = 1).value = "{}".format(j)

    def pulse_directory(self, directory, chip):
        
        print("Test--> Analyzing Pulse data for Chip {}...".format(chip))
        sys.stdout.flush()
        temp = "LN"
        dac_amps = 64
        if (temp == "RT"):
            self.vlt_slope = int_dac_fit(1)
        elif (temp == "LN"):
            self.vlt_slope = int_dac_fit(0)
        full_rows = 21
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['pulse']
        self.data_folder = self.test_folder + dir_map['data']
        
        with open(self.data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
            
        directory_file_scheme = a['Pulse_Naming2']
        data_file_scheme = a['Pulse_Naming']
        
        wb = Workbook()
        wb.remove_sheet(wb.active)
        ws_summary = wb.create_sheet()
        ws_summary.title = "Gain Summary"
        ws_summary.sheet_view.zoomScale = 70
        self.setup_spreadsheet(ws_summary, len(a["gains"]), full_rows, start = 4)
        for i, gain in enumerate(a["gains"]):
            self.gain = gain
            title = ("Gain = {}".format("{}".format(gain)))

            row = (full_rows * i) + 1
            ws_summary.merge_cells(start_row = row, start_column = 2, end_row = row, end_column = 33)
            ws_summary.cell(row = row, column = 2).value = title
            ws_summary.cell(row = row, column = 2).alignment = self.center
            ws_summary.cell(row = row, column = 2).font = self.ft
                                
            ws = wb.create_sheet()
            ws.title = title
            ws.sheet_view.zoomScale = 70
            self.setup_spreadsheet(ws, dac_amps, full_rows, start = 4)
            print ("Test--> Analyzed Gain {}".format(gain))
            
            for k, peak in enumerate(a["peaks"]):
                self.peak = peak
                title = ("Peaking Time = {}".format(peak))
                total_sections = len(a["peaks"])
                section_size = 32 // total_sections
                for dac in range(dac_amps):
                    row = (full_rows * dac) + 2
                    ws.merge_cells(start_row = row, start_column = 2 + (k * section_size), end_row = row, end_column = 1 + (k * section_size) + section_size)
                    ws.cell(row = row, column = 2 + (k * section_size)).value = title
                    ws.cell(row = row, column = 2 + (k * section_size)).alignment = self.center
                    ws.cell(row = row, column = 2 + (k * section_size)).font = self.ft
                    
                row = 2 + (i * full_rows)
                ws_summary.merge_cells(start_row = row, start_column = 2 + (k * section_size), end_row = row, end_column = 1 + (k * section_size) + section_size)  
                ws_summary.cell(row = row, column = 2 + (k * section_size)).value = title
                ws_summary.cell(row = row, column = 2 + (k * section_size)).alignment = self.center
                ws_summary.cell(row = row, column = 2 + (k * section_size)).font = self.ft
                
                for m, leak in enumerate(a["leaks"]):
                    self.leak = leak
                    title = ("Current = {}".format((leak)))
                    total_sections = len(a["leaks"])
                    section_size2 = section_size // total_sections
                    for dac in range(dac_amps):
                        row = (full_rows * dac) + 3
                        ws.merge_cells(start_row = row, start_column = 2 + (k * section_size) + (m * section_size2), 
                                       end_row = row, end_column = 1 + (k * section_size) + (m * section_size2) + section_size2)
                        ws.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).value = title
                        ws.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).alignment = self.center
                        ws.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).font = self.ft
                      
                    row = 3 + (i * full_rows)
                    ws_summary.merge_cells(start_row = row, start_column = 2 + (k * section_size) + (m * section_size2), 
                                       end_row = row, end_column = 1 + (k * section_size) + (m * section_size2) + section_size2)
                    ws_summary.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).value = title
                    ws_summary.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).alignment = self.center
                    ws_summary.cell(row = row, column = 2 + (k * section_size) + (m * section_size2)).font = self.ft
                    
                    for n, base in enumerate(a["bases"]):
                        self.base = base
                        title = ("{}".format(base))
                        for dac in range(dac_amps):
                            row = (full_rows * dac) + 4
                            ws.cell(row = row, column = (k*8) + (m*2) + n + 2).value = title
                            ws.cell(row = row, column = (k*8) + (m*2) + n + 2).alignment = self.center
                            ws.cell(row = row, column = (k*8) + (m*2) + n + 2).font = self.ft
                            self.specific_directory = directory_file_scheme.format(self.gain,self.peak,self.leak,self.base)
                            
                        row = 4 + (i * full_rows)
                        ws_summary.cell(row = row, column = (k*8) + (m*2) + n + 2).value = title
                        ws_summary.cell(row = row, column = (k*8) + (m*2) + n + 2).alignment = self.center
                        ws_summary.cell(row = row, column = (k*8) + (m*2) + n + 2).font = self.ft    
                            
                        fig = plt.figure(figsize=(16, 12), dpi=80)
                        ax = fig.add_subplot(1,1,1)
                        ax.set_xlabel("Test Charge Injection (fC)")
                        ax.set_ylabel("Pulse Height (mV)")
#                        ax.ticklabel_format(axis='x', style='sci', scilimits=(-2,2))
                        ax.set_title("Chip {} ADC Output for various injected charges".format(chip))
                        plot_path = self.data_folder + directory_file_scheme.format(self.gain,self.peak,self.leak,self.base) + "Gain_plot_Chip{}".format(chip)
                        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
                            item.set_fontsize(20)
                        ax.title.set_fontsize(30)
                        slopes = []
                        for chn in range(settings.channels):
                            full_data = []
                            for j in range(dac_amps):
                                self.dac_amp = j
                                title = ("DAC Amplitude = {}".format("{}".format(j)))
                                
                                row = (full_rows * j) + 1
                                ws.merge_cells(start_row = row, start_column = 2, end_row = row, end_column = 33)
                                ws.cell(row = row, column = 2).value = title
                                ws.cell(row = row, column = 2).alignment = self.center
                                ws.cell(row = row, column = 2).font = self.ft
            
                                self.filename = data_file_scheme.format(chn,self.gain,self.peak,self.leak,self.base, self.dac_amp)
                                row = chn + (full_rows * j) + 5
                                column = (k*8) + (m*2) + n + 2
                                cell = ws.cell(row = row , column = column)
                                pulse_height = self.pulse_file(self.data_folder + self.specific_directory + self.filename)
                                cell.value = pulse_height
                                full_data.append(pulse_height)
                                
                            slope, econstant, mV, fC = linear_fit(chip, chn, full_data, self.vlt_slope)
                            slopes.append(slope)
    
                            ax.scatter(fC, mV, marker='.')
                            ax.plot(fC, mV)
                            
                            row = (chn + 5) + (full_rows * i)
                            column = (k*8) + (m*2) + n + 2
                            cell = ws_summary.cell(row = row , column = column)
                            cell.value = slope
                            if (gain == "4.7mV"):
                                if ((slope >= settings.pulse_47_max) or (slope <= settings.pulse_47_min)):
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                                else:
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                                    
                            if (gain == "7.8mV"):
                                if ((slope >= settings.pulse_78_max) or (slope <= settings.pulse_78_min)):
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                                else:
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                                    
                            if (gain == "14mV"):
                                if ((slope >= settings.pulse_14_max) or (slope <= settings.pulse_14_min)):
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                                else:
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                                    
                            if (gain == "25mV"):
                                if ((slope >= settings.pulse_25_max) or (slope <= settings.pulse_25_min)):
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                                else:
                                    cell.font = Font(color = settings.white)
                                    cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                            
                        self.plot_stats(ax,slopes,directory_file_scheme.format(self.gain,self.peak,self.leak,self.base))
            
                        
            
                        xticks = ax.xaxis.get_major_ticks()
                        xticks[0].label1.set_visible(False)
                
                        plt.savefig (plot_path+".png")
                        plt.clf()

            wb.save(filename = self.test_folder + "{}_Pulse_Data.xlsx".format(chip))

        wb.save(filename = self.test_folder + "{}_Pulse_Data.xlsx".format(chip))
        
        print ("Test--> Baseline data analyzed for Chip {}".format(chip))
        return True
#        os.startfile(test_folder + "Baseline_Data.xlsx".format(base))
    def pulse_file(self, filepath):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(filelength/2),raw_data)

        #Ignore header for now
        data = unpacked_shorts[8:]
        data_mv = []
        for i in range(len(data)):
            data_mv.append(data[i] * settings.bits_to_mv)
            
        plots_folder = self.data_folder + self.specific_directory + "Plots/"
        if (os.path.exists(plots_folder) == False):
            os.makedirs(plots_folder)
            
        figure_data = self.plot.quickPlot(data_mv)
        ax = figure_data[0]        
        ax.set_ylabel('mV')
        ax.set_title(self.filename)
        ax.title.set_fontsize(30)
        x_lim = ax.get_xlim()
        ax.set_xlim(0, x_lim[1])
        for item in ([ax.xaxis.label, ax.yaxis.label] +ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        save_file = (plots_folder + self.filename[:len(self.filename)-4] + ".png")
        figure_data[1].savefig (save_file)
        plt.close(figure_data[1])
        
        np_data = np.array(data_mv)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=RuntimeWarning)
            pedmean = np.nanmean(np_data)
        maxmean = np.max(np_data)
    
    #Find peaks while inputting the "minimum peak height", which must be higher than half the max value found 
    #and "minimum peak distance".  Returns the index of the peaks
        peaks_index = detect_peaks(x=np_data, mph=abs((maxmean+pedmean)/2), mpd=(50)) 
        peaks_value = []
    
    #Go from index to the actual value
        for i in peaks_index :
            peaks_value.append(np_data[i])
    
    #If there were any peaks, average them
        if len(peaks_value) != 0 :
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=RuntimeWarning)
                peaksmean = np.nanmean(peaks_value[0:-1])
    
    #Or else just return the mean value of the baseline                    
        else:
            peaksmean = pedmean
        return peaksmean
        
    def plot_stats(self,plt,slope_stat,subtitle):
        subtitle_text = (.1,.94)
        average_text = (.55,.25)
        maximum_text = (.55,.20)
        minimum_text = (.55,.15)
        std_text = (.5,.10)
    
        if (slope_stat[0] > 1000):
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=RuntimeWarning)
                average = str(int(np.nanmean(slope_stat)))
            maximum = str(int(max(slope_stat)))
            minimum = str(int(min(slope_stat)))
            std_dev = str(round(np.std(slope_stat),2))
        else:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=RuntimeWarning)
                average = str(round(np.nanmean(slope_stat),2))
            maximum = str(round(max(slope_stat),2))
            minimum = str(round(min(slope_stat),2))
            std_dev = str(round(np.std(slope_stat),4))
            
        plt.annotate(subtitle, xy=subtitle_text,  xycoords='axes fraction',
        xytext=subtitle_text, textcoords='axes fraction',
        horizontalalignment='left', verticalalignment='center',
        size = 20,
        )
    
        plt.annotate("Average gain is {} mV/fC".format(average), xy=average_text,  xycoords='axes fraction',
        xytext=average_text, textcoords='axes fraction',
        horizontalalignment='left', verticalalignment='center',
        size = 20,
        )
    
        plt.annotate("Maximum gain is {} mV/fC".format(maximum), xy=maximum_text,  xycoords='axes fraction',
        xytext=maximum_text, textcoords='axes fraction',
        horizontalalignment='left', verticalalignment='center',
        size = 20,
        )
    
        plt.annotate("Minimum gain is {} mV/fC".format(minimum), xy=minimum_text,  xycoords='axes fraction',
        xytext=minimum_text, textcoords='axes fraction',
        horizontalalignment='left', verticalalignment='center',
        size = 20,
        )
    
        plt.annotate("Standard Deviation of gains is {} mV/fC".format(std_dev), xy=std_text,  xycoords='axes fraction',
        xytext=std_text, textcoords='axes fraction',
        horizontalalignment='left', verticalalignment='center',
        size = 20,
        )
        
    def DAC_directory(self, directory, chip):
        print("Test--> Analyzing DAC step data for Chip {}...".format(chip))
        sys.stdout.flush()
        #Finds whatever the name of the baseline directory is (this is so it can be changed without problem) and gets the directories within
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['dac']
        data_folder = self.test_folder + dir_map['data']
        step_folder = self.test_folder + "Plots/"
        
        try:
            os.makedirs(step_folder)
        except WindowsError:
            pass
        
        with open(data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
        data_file_scheme = a['DAC_Naming']
        
        #Create the spreadsheet, all data(noise, baseline for both baseline options) are included
        wb = Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet()
        ws.title = "DAC Steps"
        
        ws.cell(row = 1, column = 1).value = "DAC Step"
        ws.cell(row = 1, column = 2).value = "Measured Amplitude"
        ws.cell(row = 1, column = 3).value = "Difference"
        all_values = []
        for dac in range(0,64,1):
            ws.cell(row = dac+2, column = 1).value = "{}".format(dac)
            self.filename = data_file_scheme.format(dac)
            
            results = self.DAC_file(data_folder + self.filename)
            height = results[0]
            figure_data = results[1]
            ws.cell(row = dac+2, column = 2).value = height
            if (dac != 0):
                ws.cell(row = dac+2, column = 3).value = height - all_values[dac-1]
            
            save_file = (step_folder + "DAC_step{}.png".format(dac))
            figure_data[1].savefig (save_file)
            plt.close(figure_data[1])
            
            all_values.append(height)

        wb.save(filename = self.test_folder + "DAC_Step_Data.xlsx")
        
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('DAC step')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        ax.plot(all_values)
        ax.set_ylabel('mV')
        ax.set_title("DAC step Summary for Chip {}".format(chip))
        ax.title.set_fontsize(30)
        
        a = []
        for i in range(64):
            a.append(i)
        try:
            slope, constant, r_value, p_value, std_err = stats.linregress(a[3:],all_values[3:])
        except:
            slope = 0
            constant=0
        if (math.isnan(slope)):
            slope = 0
        ax.text(0.1,0.85,"Y = ({}) * X + ({})".format(slope,constant),transform=ax.transAxes, fontsize = 20)
        save_file = (self.test_folder + "{}_DAC Summary Plot.png".format(chip))
        fig_summary.savefig (save_file)
        plt.close(fig_summary) 
        lsb = height/64
        dnl_array = []
        for i in range(4,64,1):
            dnl_array.append((ws.cell(row = i + 2, column = 3).value - lsb)/lsb)
            
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('DAC step')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        ax.plot(dnl_array)
        x_lim = ax.get_xlim()
        ax.set_xlim(0, x_lim[1])
        
        response = "1 LSB is {} mV".format(lsb)
        ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
        response = "DNL(i) = (V(i) - V(i-1) - LSB) / LSB"
        ax.text(0.01,0.75,response,transform=ax.transAxes, fontsize = 20)
        ax.set_ylabel('DNL (bits)')
        ax.set_title("DNL Summary for Chip {}".format(chip))
        ax.title.set_fontsize(30)
        
        save_file = (self.test_folder + "{}_DNL Summary Plot.png".format(chip))
        fig_summary.savefig (save_file)
        plt.close(fig_summary) 
        
        inl_array = []
        for i in range(len(dnl_array)):
             if (i == 0):
                inl_array.append(0)
             else:
                inl_array.append(np.sum(dnl_array[0:i]))
            
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('DAC step')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        ax.plot(inl_array)
        x_lim = ax.get_xlim()
        ax.set_xlim(0, x_lim[1])
        lsb = height/64
        response = "1 LSB is {} mV".format(lsb)
        ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
        response = "INL(i) = DNL(0) + DNL(1) + ... + DNL(i-1)"
        ax.text(0.01,0.75,response,transform=ax.transAxes, fontsize = 20)
        ax.set_ylabel('INL (bits)')
        ax.set_title("INL Summary for Chip {}".format(chip))
        ax.title.set_fontsize(30)
        
        save_file = (self.test_folder + "{}_INL Summary Plot.png".format(chip))
        fig_summary.savefig (save_file)
        plt.close(fig_summary) 
        
        print ("Test--> DAC step data analyzed for Chip {}".format(chip))
        result = True
        return result
        
    def DAC_file(self, filepath):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(int(filelength/2)),raw_data)

        #Ignore header for now
        data = unpacked_shorts[8:]
        data_mv = []
        for i in range(len(data)):
            data_mv.append(data[i] * settings.bits_to_mv)

        peaks_index = detect_peaks(x = data_mv, mph = (settings.DAC_peak_min), mpd=settings.DAC_freq) 
        peaks_value = []
        peaks_index_fix = []
        average = np.mean(data_mv)
        maximum = max(data_mv)
        for i in peaks_index :
            if (data_mv[i] > (((maximum - average)/2) + average)):
                peaks_value.append(data_mv[i])
                peaks_index_fix.append(i)
        
        figure_data = self.plot.quickPlot(data_mv)
        ax = figure_data[0]
        ax.set_ylabel("DAC Height (mV)")       
        
        for j in peaks_index_fix:
            y_value = data_mv[j]
            ax.scatter(j/2, y_value, marker='x')
        
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        
        return [np.mean(peaks_value), figure_data]
        
    def monitor_directory(self, directory, chip):
        print("Test--> Analyzing Monitor data for Chip {}...".format(chip))
        sys.stdout.flush()
        #Finds whatever the name of the baseline directory is (this is so it can be changed without problem) and gets the directories within
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['monitor']
        data_folder = self.test_folder + dir_map['data']
        
        with open(data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
        data_file_scheme = a['Monitor_Naming']
        
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('Time (us)')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        overall_result = True
        
        fig_summary2 = plt.figure(figsize=(16, 12), dpi=80)
        ax2 = fig_summary2.add_subplot(1,1,1)
        ax2.set_xlabel('Time (us)')
        for item in ([ax2.xaxis.label, ax.yaxis.label] + ax2.get_xticklabels() + ax2.get_yticklabels()):
            item.set_fontsize(20)
        
        #Create the spreadsheet, all data(noise, baseline for both baseline options) are included
        wb = Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet()
        ws.title = "Monitor"
        ws.cell(row = 1, column = 2).value = "Result"
        ws.cell(row = 1, column = 2).alignment = self.center
        self.setup_spreadsheet(ws, 1, 1, start = 1)
        
        baseline_file = directory + dir_map['baseline_rms'] + "{}_Baseline_Data.xlsx".format(chip)
        wb_ref = load_workbook(filename = baseline_file, read_only = True)
        ws_ref = wb_ref['200mV_mean']
        
        peaks = [] 
        for chn in range(settings.channels):
            self.filename = data_file_scheme.format(chn)
            row = chn + 2
            cell = ws.cell(row = row , column = 2)
            
            baseline_row = chn + 5
            baseline = ws_ref.cell(row = baseline_row , column = 2).value
            
            sample_pulse, result, average, corrected_pulse = self.monitor_file(data_folder + self.filename, cell, baseline)
            ws.cell(row = row , column = 3).value = average
            overall_result = result and overall_result
            peaks.append(average)
            ax.plot(sample_pulse)
            ax2.plot(corrected_pulse)
                    
        ax.set_ylabel('mV')
        ax.set_title("Sample pulses for Chip {} Monitor Test".format(chip))
        ax.title.set_fontsize(30)
        save_file = (self.test_folder + "Sample_Pulses.png")
        fig_summary.savefig (save_file)
        plt.close(fig_summary)

        
        average_peak = np.mean(peaks)
        ws.cell(row = 18, column = 1).value = "Average"
        ws.cell(row = 18, column = 3).value = average_peak
        ws.cell(row = 1, column = 3).value = "Peak"
        ws.cell(row = 1, column = 3).alignment = self.center
        ws.cell(row = 1, column = 4).value = "Difference From Average"
        ws.cell(row = 1, column = 4).alignment = self.center
        ws.cell(row = 1, column = 5).value = "Percent From Average"
        ws.cell(row = 1, column = 4).alignment = self.center
        differences = []
        for chn in range(settings.channels):
            row = chn + 2
            difference = average_peak - (ws.cell(row = row , column = 3).value)
            differences.append(difference)
            ws.cell(row = chn + 2, column = 4).value = difference
            ws.cell(row = chn + 2, column = 4).alignment = self.center
            ws.cell(row = chn + 2, column = 5).value = "{:.2f}%".format((difference / average_peak) * 100)
            ws.cell(row = chn + 2, column = 5).alignment = self.center
            
            
        ax2.set_ylabel('mV')
        ax2.set_title("Sample pulses for Chip {} Monitor Test, baseline corrected".format(chip))
        ax2.title.set_fontsize(30)
        response1 = "Average peak was {:.2f}".format(average_peak)
        response2 = "Average deviation was {:.2f}%".format(np.mean(differences))
        response3 = "Maximum deviation was {:.2f}%".format((max(differences)/average_peak) * 100)
        ax2.text(0.01,0.8,response1,transform=ax.transAxes, fontsize = 20)
        ax2.text(0.01,0.75,response2,transform=ax.transAxes, fontsize = 20)
        ax2.text(0.01,0.7,response3,transform=ax.transAxes, fontsize = 20)
        
        save_file = (self.test_folder + "Sample_Pulses_Gain_Matched.png")
        fig_summary2.savefig (save_file)
        plt.close(fig_summary2)      
        
        wb.save(filename = self.test_folder + "{}_Monitor_Data.xlsx".format(chip))
        
        print ("Test--> Monitor data analyzed for Chip {}".format(chip))
        return overall_result
                    
    def monitor_file(self, filepath, cell, baseline):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(int(filelength/2)),raw_data)

        #Ignore header for now
#        data = unpacked_shorts[8:]
        data = unpacked_shorts
        data_mv = []
        for i in range(len(data)):
            data_mv.append(data[i] * settings.bits_to_mv)
        
        failure = False
        peaks_index = detect_peaks(x = data_mv, mph = (settings.monitor_peak_min / 2), mpd=50) 
        peaks_value = []
        peaks_index_fix = []
        for i in peaks_index:
            if (data_mv[i] > settings.monitor_peak_min):
                peaks_value.append(data_mv[i])
                peaks_index_fix.append(i)
#        print ("Chip {}, Channel {} has peak values {}".format(chip, chn, peaks_value))
        #Check if the peak is in the wrong place (happens when it's not synced)
        response = "Good response, channel is detected"
        
        peak_num = len(peaks_index_fix)
        if ((peak_num <= settings.monitor_peaks_min) or (peak_num >= settings.monitor_peaks_max)):
            failure = "num"
            response = "Number of peaks should be between {} and {}, but it was {:.0f}".format(settings.monitor_peaks_min, settings.monitor_peaks_max, peak_num)

#        if (failure != False):
        rejects_folder = self.test_folder + "Plots/"
        if (os.path.exists(rejects_folder) == False):
            os.makedirs(rejects_folder)

        figure_data = self.plot.quickPlot(data_mv)
        ax = figure_data[0]
        for j in peaks_index_fix:
            y_value = data_mv[j]
            ax.scatter(j/2, y_value, marker='x')
            
        ax.set_ylabel('mV')
        ax.set_title(self.filename)
        ax.title.set_fontsize(30)
        ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
        for item in ([ax.xaxis.label, ax.yaxis.label] +ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        save_file = (rejects_folder + self.filename[:len(self.filename)-4] + ".png")
        figure_data[1].savefig (save_file)
        plt.close(figure_data[1])
            
        if (failure == False):
            cell.value = "Pass"
            cell.font = Font(color = settings.white)
            cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
        else:
            cell.value = "Fail"
            cell.font = Font(color = settings.white)
            cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                
        plot_data = data_mv[500-settings.alive_plot_x_back : 500 + settings.alive_plot_x_forward]
        if (len(peaks_index_fix) >= 2):
            peak = peaks_index_fix[1]
            plot_data = data_mv[peak - settings.monitor_plot_x_back : peak + settings.monitor_plot_x_forward]

        data_mv = []
        for i in range(len(data)):
            data_mv.append((data[i] * settings.bits_to_mv) - baseline)
        
        peaks_index = detect_peaks(x = data_mv, mph = (settings.monitor_peak_min / 2), mpd=settings.monitor_freq - 50) 
        peaks_value = []
        peaks_index_fix = []
        for i in peaks_index:
            if (data_mv[i] > settings.monitor_peak_min):
                peaks_value.append(data_mv[i])
                peaks_index_fix.append(i)
                
        plot_data_corrected = data_mv[500-settings.alive_plot_x_back : 500 + settings.alive_plot_x_forward]
        if (len(peaks_index_fix) >= 2):
            peak = peaks_index_fix[1]
            plot_data_corrected = data_mv[peak - settings.monitor_plot_x_back : peak + settings.monitor_plot_x_forward]
            
        average = np.mean(peaks_value)
            
        return plot_data, not failure, average, plot_data_corrected
        
#__INIT__#
    def __init__(self):
        self.test_folder = None
        self.filename = None
        self.curr_analysis = None
        self.plot = plot_functions()
        self.center = Alignment(horizontal='center')
        self.ft = Font(bold=True)
        self.analyze3 = Data_Analysis3()
        
if __name__ == "__main__":
    single_channels = []
    for i in range (16):
        single_channels.append("D:\Eric\Quad_Data_2017_09_26\\axestest\ch{}.dat".format(i))
    
    Data_Analysis().PlotSingles(single_channels)
    #Data_Analysis().UnpackData("D:\\nEXO\\2017_06_19\\" + "ped.dat")
    #Data_Analysis().Missing_Packet_Check("D:\\Eric\\Packets\\")
    #Data_Analysis().Seperate_Packets("D:\\Eric\\Packets\\", 4, 4)
