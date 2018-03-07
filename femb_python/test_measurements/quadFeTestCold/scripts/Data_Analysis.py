# -*- coding: utf-8 -*-
"""
Created on Fri Jun  2 11:50:17 2017

@author: vlsilab2
"""

import os
import sys
import struct
from datetime import datetime
import matplotlib.pyplot as plt
plt.switch_backend('agg')
import pickle
import numpy as np
from femb_python.test_measurements.quadFeTestCold.user_settings import user_editable_settings
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
import matplotlib.patches as mpatches
from femb_python.test_measurements.quadFeTestCold.scripts.detect_peaks import detect_peaks
from femb_python.test_measurements.quadFeTestCold.scripts.plotting import plot_functions
settings = user_editable_settings()
import warnings
with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=RuntimeWarning)

class Data_Analysis:
    
    #Puts the channel designations where they should be, as long as you know how many discrete sections there are and what the spacing between them is
    def setup_spreadsheet(self, ws, sections, rows, start):
        for i in range(sections):
            ws.cell(row = (rows * i) + start, column = 1).value = "Channel"
            for j in range(16):
                ws.cell(row = j + (rows * i) + start + 1, column = 1).value = "{}".format(j)
            
    def baseline_directory(self, directory, chip):
        print("Test--> Analyzing Baseline data for Chip {}...".format(chip))
        sys.stdout.flush()
        full_rows = 21
        #Finds whatever the name of the baseline directory is (this is so it can be changed without problem) and gets the directories within
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['baseline_rms']
        data_folder = self.test_folder + dir_map['data']
        
        with open(data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
        data_file_scheme = a['Baseline_Naming']
        
        #Create the spreadsheet, all data(noise, baseline for both baseline options) are included
        wb = Workbook()
        wb.remove(wb.active)
        
        sections = len(a["gains"])
#        ws_tuple = ["_mean","_std"]
        ws_tuple = ["_mean"]
        overall_result = True
        
        for ws_name in ws_tuple:
            #But each analysis gets a different summary plot
            self.curr_analysis = ws_name
            fig_summary = plt.figure(figsize=(16, 12), dpi=80)
            ax = fig_summary.add_subplot(1,1,1)
            ax.set_xlabel('Channel')

            for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
                item.set_fontsize(20)
            
            for num, base in enumerate(a["bases"]):
                #Each base gets a new sheet, and zooming out lets you see everything as soon as you open it
                self.base = base
                ws = wb.create_sheet()
                ws.title = base + ws_name
                ws.sheet_view.zoomScale = 70
                self.setup_spreadsheet(ws, sections, full_rows, start = 4)
                
                for i, gain in enumerate(a["gains"]):
                    self.gain = gain
                    title = ("Gain = {}".format("{}/fC".format(gain)))
                    row = (full_rows * i) + 1
                    section_size = len(a["peaks"]) * len(a["leaks"]) * len(a["buffs"])
                    ws.merge_cells(start_row = row, start_column = 2, end_row = row, end_column = 1 + section_size)
                    ws.cell(row = row, column = 2).value = title
                    ws.cell(row = row, column = 2).alignment = self.center
                    ws.cell(row = row, column = 2).font = self.ft
                    
                    for j, peak in enumerate(a["peaks"]):
                        title = ("Peaking Time = {}".format(peak))
                        row = (full_rows * i) + 2
                        
                        section_size = len(a["leaks"]) * len(a["buffs"])
                        
                        ws.merge_cells(start_row = row, start_column = 2 + (j * section_size), end_row = row, end_column = 1 + (j * section_size) + section_size)
                        ws.cell(row = row, column = 2 + (j * section_size)).value = title
                        ws.cell(row = row, column = 2 + (j * section_size)).alignment = self.center
                        ws.cell(row = row, column = 2 + (j * section_size)).font = self.ft
                        
                        for k, leak in enumerate(a["leaks"]):
                            title = ("{}".format((leak)))
                            row = (full_rows * i) + 3
                            
                            total_sections = len(a["leaks"])
                            #divides the section_size of the previous for loop
                            section_size2 = section_size // total_sections
                            ws.merge_cells(start_row = row, start_column = 2 + (j * section_size) + (k * section_size2), 
                                           end_row = row, end_column = 1 + (j * section_size) + (k * section_size2) + section_size2)
                            ws.cell(row = row, column = 2 + (j * section_size) + (k * section_size2)).value = title
                            ws.cell(row = row, column = 2 + (j * section_size) + (k * section_size2)).alignment = self.center
                            ws.cell(row = row, column = 2 + (j * section_size) + (k * section_size2)).font = self.ft
                    
                            for m, buff in enumerate(a["buffs"]):
                                title = ("{}".format(buff))
                                row = (full_rows * i) + 4
                                ws.cell(row = row, column = (j*8) + (k*2) + m + 2).value = title
                                ws.cell(row = row, column = (j*8) + (k*2) + m + 2).alignment = self.center
                                ws.cell(row = row, column = (j*8) + (k*2) + m + 2).font = self.ft
                                
                                plot_color = (m * 128) + ((k * 64) << 8) + ((j * 64) << 16) + ((i * 64) << 24) + ((num * 128) << 32)
                                plot_color = "#{:x}{:x}{:x}{:x}{:x}{:x}".format(((plot_color >> 20) & 0xF),((plot_color >> 16) & 0xF),((plot_color >> 12) & 0xF),
                                ((plot_color >> 8) & 0xF),((plot_color >> 4) & 0xF),((plot_color) & 0xF))
                                
#                                if m == 0:
#                                    plot_color = "black"
#                                if m == 1:
#                                    plot_color = "red"
                                
                                for chn in range(settings.channels):
                                    self.filename = data_file_scheme.format(chn,gain,peak,leak,buff,self.base)
                                    row = chn + (full_rows * i) + 5
                                    column = (j*8) + (k*2) + m + 2
                                    cell = ws.cell(row = row , column = column)
                                    
                                    #To make sure you don't plot the noise 4.7 and 7.8 gains, they're usually bad
                                    if ((ws_name == "_mean") or (gain == "14mV") or (gain == "25mV")):
                                        self.print_if_outside = True
                                    else:
                                        self.print_if_outside = False
                                        
                                    y, result = self.baseline_file(data_folder + self.filename, cell)
                                    overall_result = overall_result and result
                                    
                                    if (self.print_if_outside == True):
                                        ax.scatter(chn, y, color=plot_color)
                #average the baselines
                baselines = []
                for i in range(len(a["gains"])):
                    for j in range(2,34,1):
                        for chn in range(settings.channels):
                            val = ws.cell(row = chn + (full_rows * i) + 5, column = j).value
                            if (val != None):
                                baselines.append(val)

                average = np.mean(baselines)
                if (ws_name == "_mean"):
                    ax.set_ylabel('mV')
                    ax.set_title("Baseline Summary for Chip {}".format(chip))
                    if (base == "200mV"):
                        y_pos = 0.4
                        comment = "Average 200 mV baseline is {:.2f} mV".format(average)
                    elif (base == "900mV"):
                        y_pos = 0.6
                        comment = "Average 900 mV baseline is {:.2f} mV".format(average)
                        
                    ax.text(0.01,y_pos,comment,transform=ax.transAxes, fontsize = 20)
                elif (ws_name == "_std"):
                    if (base == "200mV"):
                        comment = "Average noise is {:.2f} electrons".format(np.mean(average))
                    elif (base == "900mV"):
                        comment = "Average noise is {:.2f} electrons".format(np.mean(average))
                        
                    ax.set_ylabel('electrons')
                    ax.set_title("Noise Summary for Chip {}, baseline {}".format(chip, base))
                    y_pos = 0.01
                    ax.text(0.01,y_pos,comment,transform=ax.transAxes, fontsize = 20)
                        
                    chan = range(-1, 17)
                    ax.set_xticks(chan)
                    y_lim = ax.get_ylim()
                    ax.set_ylim(1, y_lim[1])
                    xticks = ax.xaxis.get_major_ticks()
                    xticks[0].set_visible(False)
                    xticks[17].set_visible(False)
                    ax.title.set_fontsize(30)

                    ax.axhspan(0, settings.noise_reject_min, color='red', alpha=settings.alpha, lw=0)
                    ax.axhspan(settings.noise_reject_min, settings.noise_good_min, color='yellow', alpha=settings.alpha)
                    ax.axhspan(settings.noise_good_min, settings.noise_good_max, color='green', alpha=settings.alpha, lw = 0)
                    ax.axhspan(settings.noise_good_max, settings.noise_reject_max, color='yellow', alpha=settings.alpha)
                    ax.axhspan(settings.noise_reject_max, y_lim[1], color='red', alpha=settings.alpha)

                    save_file = (self.test_folder + "{}_Noise Summary Plot_{}.png".format(chip, base))
                    fig_summary.savefig (save_file)
                    plt.close(fig_summary)
                    fig_summary = plt.figure(figsize=(16, 12), dpi=80)
                    ax = fig_summary.add_subplot(1,1,1)
                    ax.set_xlabel('Channel')
                    for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
                        item.set_fontsize(20)
            
            
            
            if (ws_name == "_mean"):
                chan = range(-1, 17)
                ax.set_xticks(chan)
                y_lim = ax.get_ylim()
                ax.set_ylim(1, y_lim[1])
                xticks = ax.xaxis.get_major_ticks()
                xticks[0].set_visible(False)
                xticks[17].set_visible(False)
                ax.title.set_fontsize(30)
            
                ax.axhspan(0, settings.baseline_200_reject_min, color='red', alpha=settings.alpha, lw=0)
                ax.axhspan(settings.baseline_200_reject_min, settings.baseline_200_good_min, color='yellow', alpha=settings.alpha)
                ax.axhspan(settings.baseline_200_good_min, settings.baseline_200_good_max, color='green', alpha=settings.alpha, lw = 0)
                
                ax.axhspan(settings.baseline_200_good_max, settings.baseline_200_reject_max, color='yellow', alpha=settings.alpha)
                ax.axhspan(settings.baseline_200_reject_max, settings.baseline_900_reject_min, color='red', alpha=settings.alpha)
                ax.axhspan(settings.baseline_900_reject_min, settings.baseline_900_good_min, color='yellow', alpha=settings.alpha)
                
                ax.axhspan(settings.baseline_900_good_min, settings.baseline_900_good_max, color='green', alpha=settings.alpha)
                ax.axhspan(settings.baseline_900_good_max, settings.baseline_900_reject_max, color='yellow', alpha=settings.alpha)
                ax.axhspan(settings.baseline_900_reject_max, y_lim[1], color='red', alpha=settings.alpha)
                save_file = (self.test_folder + "{}_Baseline Summary Plot.png".format(chip))
                
                comment = "Gain = 14mV/fC, Peaking Time = 2us, Leakage = 500 pA, Buffer on"
                ax.text(0.01,0.5,comment,transform=ax.transAxes, fontsize = 20)
#                black_patch = mpatches.Patch(color='black', label='Buffer Off')
#                blue_patch = mpatches.Patch(color='red', label='Buffer On')
#                plt.legend(handles=[black_patch, blue_patch])
            
                fig_summary.savefig (save_file)
                plt.close(fig_summary)
        wb.save(filename = self.test_folder + "{}_Baseline_Data.xlsx".format(chip))
        
        print ("Test--> Baseline data analyzed for Chip {}".format(chip))
        return overall_result
#        os.startfile(test_folder + "Baseline_Data.xlsx".format(base))
    def baseline_file(self, filepath, cell):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(int(filelength/2)),raw_data)

        #Ignore header for now
        data = unpacked_shorts[8:]
        data_mv = []
        for i in range(len(data)):
            data_mv.append(data[i] * settings.bits_to_mv)
        
        if (self.curr_analysis == "_mean"):
            mean = np.mean(data_mv)
            cell.value = mean
            response = self.baseline_analysis(cell = cell)
            
        elif (self.curr_analysis == "_std"):
            std = np.std(data_mv) / float(self.gain[:self.gain.find("m")]) / (1.6E-19) * (1E-15)
            std_mv = np.std(data_mv)
            cell.value = std
            response = self.noise_analysis(cell = cell, mv = std_mv)
            
        if ((response != True) and (self.print_if_outside == True)):
            rejects_folder = self.test_folder + "Outside Range/"
            if (os.path.exists(rejects_folder) == False):
                os.makedirs(rejects_folder)

            figure_data = self.plot.quickPlot(data_mv)
            ax = figure_data[0]
            ax.set_ylabel('mV')
            ax.set_title(self.filename)
            ax.title.set_fontsize(30)
            ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
            response = False
            for item in ([ax.xaxis.label, ax.yaxis.label] +ax.get_xticklabels() + ax.get_yticklabels()):
                item.set_fontsize(20)
            save_file = (rejects_folder + self.filename[:len(self.filename)-4] + self.curr_analysis + ".png")
            figure_data[1].savefig (save_file)
            plt.close(figure_data[1])
                
        return cell.value, response
        
    def baseline_analysis(self, cell):
        value = int(cell.value)
        if (self.base == "200mV"):
            if ((value >= settings.baseline_200_reject_max) or (value <= settings.baseline_200_reject_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                return ("Baseline should be between {} and {}, but it was {:.2f} ".format(settings.baseline_200_reject_min, 
                        settings.baseline_200_reject_max, value))
            if ((value >= settings.baseline_200_acceptable_max) or (value <= settings.baseline_200_acceptable_min)):
                cell.fill = PatternFill(start_color = settings.yellow, end_color = settings.yellow, fill_type = 'solid')
                return ("Baseline should be between {} and {}, but it was {:.2f} ".format(settings.baseline_200_acceptable_min, 
                        settings.baseline_200_acceptable_max, value))
            if ((value <= settings.baseline_200_good_max) or (value >= settings.baseline_200_good_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                return True
#            return True
                
        if (self.base == "900mV"):
            if ((value >= settings.baseline_900_reject_max) or (value <= settings.baseline_900_reject_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                return ("Baseline should be between {} and {}, but it was {:.2f} ".format(settings.baseline_900_reject_min, 
                        settings.baseline_900_reject_max, value))
            if ((value >= settings.baseline_900_acceptable_max) or (value <= settings.baseline_900_acceptable_min)):
                cell.fill = PatternFill(start_color = settings.yellow, end_color = settings.yellow, fill_type = 'solid')
                return ("Baseline should be between {} and {}, but it was {:.2f} ".format(settings.baseline_900_acceptable_min, 
                        settings.baseline_900_acceptable_max, value))
            if ((value <= settings.baseline_900_good_max) or (value >= settings.baseline_900_good_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
                return True
#            return True
            
    def noise_analysis(self, cell, mv):
        value = int(cell.value)
        #We only care about 14 mv or 25 mv case case
        if ((self.gain == "14mV") or (self.gain == "25mV")):
            if ((value >= settings.noise_reject_max) or (value <= settings.noise_reject_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                return ("Noise should be between {} and {}, but it was {:.2f} electrons ({:.2f} mV)".format(settings.noise_reject_min, 
                        settings.noise_reject_max, value, mv))
            if ((value >= settings.noise_acceptable_max) or (value <= settings.noise_acceptable_min)):
                cell.fill = PatternFill(start_color = settings.yellow, end_color = settings.yellow, fill_type = 'solid')
                return ("Noise should be between {} and {}, but it was {:.2f} electrons ({:.2f} mV)".format(settings.noise_acceptable_min, 
                        settings.noise_acceptable_max, value, mv))
            if ((value <= settings.noise_good_max) or (value >= settings.noise_good_min)):
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
#                return True
            
        return True
    def alive_directory(self, directory, chip):
        print("Test--> Analyzing 'Channel Alive' data for Chip {}...".format(chip))
        sys.stdout.flush()
        full_rows = 18
        sections = 1
        #Finds whatever the name of the baseline directory is (this is so it can be changed without problem) and gets the directories within
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['alive']
        data_folder = self.test_folder + dir_map['data']
        
        with open(data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
        data_file_scheme = a['Alive_Naming']
        
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('Time (us)')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
            
        overall_result = True
        #Create the spreadsheet, all data(noise, baseline for both baseline options) are included
        wb = Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet()
        ws.title = "Channel Alive"
        self.setup_spreadsheet(ws, sections, full_rows, start = 2)
        cols_per_setting = len(a["leaks"])
        for i, test in enumerate(a["tests"]):
            title = ("Test = {}".format(test))
            row = 1
            
            ws.merge_cells(start_row = row, start_column = (cols_per_setting * i) + 2, end_row = row, end_column = (cols_per_setting * i) + (len(a["leaks"]) + 1))
            ws.cell(row = row, column = (cols_per_setting * i) + 2).value = title
            ws.cell(row = row, column = (cols_per_setting * i) + 2).alignment = self.center
            for j, leak in enumerate(a["leaks"]):
                title = ("Leak = {}".format(leak))
                row = 2
                ws.cell(row = row, column = (i * cols_per_setting) + j + 2).value = title
                ws.cell(row = row, column = (i * cols_per_setting) + j + 2).alignment = self.center
                
                for chn in range(settings.channels):
                    self.filename = data_file_scheme.format(chn,leak,test)
                    row = chn + 3
                    column = (i * cols_per_setting) + j + 2
                    cell = ws.cell(row = row , column = column)
                    sample_pulse, result = self.alive_file(data_folder + self.filename, cell)
                    overall_result = overall_result and result
                    if (test == "test_off"):
                        plot_color = "green"
                    elif (test == "test_ext"):
                        plot_color = "red"
                    ax.plot(sample_pulse, color = plot_color)
        ws.cell(row = settings.channels + 6, column = 1).value = "Power Cycles"
        data_file_scheme = a['Alive_Naming2']
        for cycle in range(settings.power_cycles):
            ws.cell(row = settings.channels + 7 + cycle, column = 1).value = cycle + 1
            self.filename = data_file_scheme.format(chn,leak,test,cycle)
            row = cycle + 7 + settings.channels
            cell = ws.cell(row = row , column = 2)
            sample_pulse, result = self.alive_file(data_folder + self.filename, cell)
            overall_result = overall_result and result
            plot_color = "blue"
            ax.plot(sample_pulse, color = plot_color)
                    
        green_patch = mpatches.Patch(color='green', label='Input pin to FE')
        red_patch = mpatches.Patch(color='red', label='Test pin to FE')
        blue_patch = mpatches.Patch(color='blue', label='Pulses after power cycles')
        plt.legend(handles=[green_patch, red_patch, blue_patch])
                    
        ax.set_ylabel('mV')
        ax.set_title("Sample pulses for Chip {} Channel Alive Test".format(chip))
        ax.title.set_fontsize(30)
        save_file = (self.test_folder + "Sample_Pulses.png")
        fig_summary.savefig (save_file)
        plt.close(fig_summary)      
        
        wb.save(filename = self.test_folder + "{}_Pulse_Alive_Data.xlsx".format(chip))
        
        print ("Test--> 'Channel Alive' data analyzed for Chip {}".format(chip))
        return overall_result
                    
    def alive_file(self, filepath, cell):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(int(filelength/2)),raw_data)

        #Ignore header for now
        data = unpacked_shorts[8:]
        data_mv = []
        for i in range(len(data)):
            data_mv.append(data[i] * settings.bits_to_mv)
        
        failure = False
        peaks_index = detect_peaks(x = data_mv, mph = (settings.test_peak_min / 4), mpd=settings.test_TP_Period - 50) 
        peaks_value = []
        peaks_index_fix = []
        maximum = max(data_mv)
        for i in peaks_index :
            if (data_mv[i] > settings.test_peak_min) and (data_mv[i] > (maximum - settings.under_max)):
                peaks_value.append(data_mv[i])
                peaks_index_fix.append(i)
                
#        print ("Chip {}, Channel {} has peak values {}".format(chip, chn, peaks_value))
        #Check if the peak is in the wrong place (happens when it's not synced)
        response = "Good response, channel is alive"
        
        peak_num = len(peaks_index_fix)
        if ((peak_num <= settings.test_peaks_min) or (peak_num >= settings.test_peaks_max)):
            failure = "num"
            response = "Number of peaks should be between {} and {}, but it was {:.0f}.  Wanted peaks between {} and {}".format(settings.test_peaks_min,settings.test_peaks_max, peak_num, settings.test_peak_min, maximum - settings.under_max)

#        if (failure != False):
        rejects_folder = self.test_folder + "Plots/"
        if (os.path.exists(rejects_folder) == False):
            os.makedirs(rejects_folder)

        figure_data = self.plot.quickPlot(data_mv)
        ax = figure_data[0]
        for j in peaks_index_fix:
            y_value = data_mv[j]
            ax.scatter(j/2, y_value, marker='x')
        
        ax.set_ylabel('mV')
        ax.set_title(self.filename)
        ax.title.set_fontsize(30)
        x_lim = ax.get_xlim()
        ax.set_xlim(0, x_lim[1])
        ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
        for item in ([ax.xaxis.label, ax.yaxis.label] +ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        save_file = (rejects_folder + self.filename[:len(self.filename)-4] + ".png")
        figure_data[1].savefig (save_file)
        plt.close(figure_data[1])
            
        if (failure == False):
            cell.value = "Pass"
            cell.font = Font(color = settings.white)
            cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
        else:
            cell.value = "Fail"
            cell.font = Font(color = settings.white)
            cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                
        plot_data = data_mv[500-settings.alive_plot_x_back : 500 + settings.alive_plot_x_forward]
        if (len(peaks_index_fix) >= 3):
            peak = peaks_index_fix[1]
            plot_data = data_mv[peak - settings.alive_plot_x_back : peak + settings.alive_plot_x_forward]
            
        return plot_data, not failure

#__INIT__#
    def __init__(self):
        self.test_folder = None
        self.filename = None
        self.curr_analysis = None
        self.plot = plot_functions()
        self.center = Alignment(horizontal='center')
        self.ft = Font(bold=True)
        
if __name__ == "__main__":
    single_channels = []
    for i in range (16):
        single_channels.append("D:\Eric\Quad_Data_2017_09_26\\axestest\ch{}.dat".format(i))
    
    Data_Analysis().PlotSingles(single_channels)
    #Data_Analysis().UnpackData("D:\\nEXO\\2017_06_19\\" + "ped.dat")
    #Data_Analysis().Missing_Packet_Check("D:\\Eric\\Packets\\")
    #Data_Analysis().Seperate_Packets("D:\\Eric\\Packets\\", 4, 4)
