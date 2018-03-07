# -*- coding: utf-8 -*-
"""
Created on Fri Jun  2 11:50:17 2017

@author: vlsilab2
"""

import os
import sys
import struct
import matplotlib.pyplot as plt
plt.switch_backend('agg')
import pickle
import numpy as np
from femb_python.test_measurements.quadFeTestCold.user_settings import user_editable_settings
from openpyxl import Workbook, load_workbook
import warnings
from femb_python.test_measurements.quadFeTestCold.scripts.int_dac_fit import int_dac_fit
from femb_python.test_measurements.quadFeTestCold.scripts.linear_fit_m import linear_fit
import matplotlib.patches as mpatches
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from femb_python.test_measurements.quadFeTestCold.scripts.detect_peaks import detect_peaks
from femb_python.test_measurements.quadFeTestCold.scripts.plotting import plot_functions
from scipy import stats
import math
settings = user_editable_settings()
import warnings
with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=RuntimeWarning)

class Data_Analysis3:
    
    #Puts the channel designations where they should be, as long as you know how many discrete sections there are and what the spacing between them is
    def setup_spreadsheet(self, ws, sections, rows, start):
        for i in range(sections):
            ws.cell(row = (rows * i) + start, column = 1).value = "Channel"
            for j in range(16):
                ws.cell(row = j + (rows * i) + start + 1, column = 1).value = "{}".format(j)

    def gain_match_directory(self, directory, chip, wb):
        print("Test--> Analyzing gain match data for Chip {}...".format(chip))
        sys.stdout.flush()
        #Finds whatever the name of the baseline directory is (this is so it can be changed without problem) and gets the directories within
        with open(directory + 'directory_map.cfg', 'rb') as f:
            dir_map = pickle.load(f)
        self.test_folder = directory + dir_map['monitor']
        data_folder = self.test_folder + dir_map['data']
        
        baseline_file = directory + dir_map['baseline_rms'] + "{}_Baseline_Data.xlsx".format(chip)
        wb_ref = load_workbook(filename = baseline_file, read_only = True)
        ws_ref = wb_ref['200mV_mean']

        with open(data_folder + 'Configuration.cfg', 'rb') as f:
            a = pickle.load(f)
        data_file_scheme = a['Monitor_Naming']
        
        fig_summary = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig_summary.add_subplot(1,1,1)
        ax.set_xlabel('Time (us)')
        for item in ([ax.xaxis.label, ax.yaxis.label] + ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
            
        ws = wb.get_sheet_by_name("Monitor")
        ws.cell(row = 1, column = 3).value = "Peak"
        ws.cell(row = 1, column = 3).alignment = self.center
        ws.cell(row = 1, column = 4).value = "Difference From Average"
        ws.cell(row = 1, column = 4).alignment = self.center
        ws.cell(row = 1, column = 5).value = "Percent From Average"
        ws.cell(row = 1, column = 5).alignment = self.center
        self.setup_spreadsheet(ws, 1, 1, start = 1)
        
        peaks = []        
        for chn in range(settings.channels):
            row = chn + 5
            baseline = ws_ref.cell(row = row , column = 1).value      
            self.filename = data_file_scheme.format(chn)
            row = chn + 2
            sample_pulse, result = self.gain_match_file(data_folder + self.filename, baseline)
            ws.cell(row = row , column = 3).value = result
            peaks.append(result)
            ax.plot(sample_pulse)
            
        average_peak = np.mean(peaks)
        ws.cell(row = 18, column = 1).value = "Average"
        ws.cell(row = 18, column = 3).value = average_peak
        differences = []
        for chn in range(settings.channels):
            row = chn + 2
            difference = average_peak - (ws.cell(row = row , column = 3).value)
            differences.append(difference)
            ws.cell(row = chn + 2, column = 4).value = difference
            ws.cell(row = chn + 2, column = 4).alignment = self.center
            ws.cell(row = chn + 2, column = 5).value = "{:.2f}%".format((difference / average_peak) * 100)
            ws.cell(row = chn + 2, column = 5).alignment = self.center
            
            
        ax.set_ylabel('mV')
        ax.set_title("Sample pulses for Chip {} Monitor Test, baseline corrected".format(chip))
        ax.title.set_fontsize(30)
        
        response1 = "Average peak was {:.2f}".format(average_peak)
        response2 = "Average deviation was {:.2f}%".format(np.mean(differences))
        response3 = "Maximum deviation was {:.2f}%".format((max(differences)/average_peak) * 100)
        ax.text(0.01,0.95,response1,transform=ax.transAxes, fontsize = 20)
        ax.text(0.01,0.85,response2,transform=ax.transAxes, fontsize = 20)
        ax.text(0.01,0.75,response3,transform=ax.transAxes, fontsize = 20)
        
        save_file = (self.test_folder + "Sample_Pulses_Gain_Matched.png")
        fig_summary.savefig (save_file)
        plt.close(fig_summary)      
        
        wb.save(filename = self.test_folder + "{}_Gain_Match_Data.xlsx".format(chip))
        
        print ("Test--> Monitor data analyzed for Chip {}".format(chip))
        return wb
                    
    def gain_match_file(self, filepath, baseline):
        fileinfo  = os.stat(filepath)
        filelength = fileinfo.st_size

        with open(filepath, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        #Filelength is in bytes, but the data we're getting is 2 bytes each
        unpacked_shorts = struct.unpack_from(">{}H".format(filelength/2),raw_data)

        #Ignore header for now
#        data = unpacked_shorts[8:]
        data = unpacked_shorts
        data_mv = []
        print (baseline)
        print (type(baseline))
        for i in range(len(data)):
            data_mv.append((data[i] * settings.bits_to_mv) - baseline)
        
        peaks_index = detect_peaks(x = data_mv, mph = (settings.monitor_peak_min / 2), mpd=settings.monitor_freq - 50) 
        peaks_value = []
        peaks_index_fix = []
        for i in peaks_index:
            if (data_mv[i] > settings.monitor_peak_min):
                peaks_value.append(data_mv[i])
                peaks_index_fix.append(i)
#        print ("Chip {}, Channel {} has peak values {}".format(chip, chn, peaks_value))
        #Check if the peak is in the wrong place (happens when it's not synced)
        response = "Good response, channel is detected"
        
        peak_num = len(peaks_index_fix)
        if ((peak_num <= settings.monitor_peaks_min) or (peak_num >= settings.monitor_peaks_max)):
            response = "Number of peaks should be between {} and {}, but it was {:.0f}".format(settings.monitor_peaks_min, settings.monitor_peaks_max, peak_num)

#        if (failure != False):
        rejects_folder = self.test_folder + "Plots\\"
        if (os.path.exists(rejects_folder) == False):
            os.makedirs(rejects_folder)

        figure_data = self.plot.quickPlot(data_mv)
        ax = figure_data[0]
        for j in peaks_index_fix:
            y_value = data_mv[j]
            ax.scatter(j/2, y_value, marker='x')
            
        ax.set_ylabel('mV')
        ax.set_title(self.filename)
        ax.title.set_fontsize(30)
        ax.text(0.01,0.95,response,transform=ax.transAxes, fontsize = 20)
        for item in ([ax.xaxis.label, ax.yaxis.label] +ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(20)
        save_file = (rejects_folder + self.filename[:len(self.filename)-4] + "_Gain_Match.png")
        figure_data[1].savefig (save_file)
        plt.close(figure_data[1])
                
        plot_data = data_mv[500-settings.alive_plot_x_back : 500 + settings.alive_plot_x_forward]
        if (len(peaks_index_fix) >= 2):
            peak = peaks_index_fix[1]
            plot_data = data_mv[peak - settings.monitor_plot_x_back : peak + settings.monitor_plot_x_forward]
            
        average = np.mean(peaks_value)
            
        return plot_data, average
        
#__INIT__#
    def __init__(self):
        self.test_folder = None
        self.filename = None
        self.curr_analysis = None
        self.plot = plot_functions()
        self.center = Alignment(horizontal='center')
        self.ft = Font(bold=True)
        
if __name__ == "__main__":
    single_channels = []
    for i in range (16):
        single_channels.append("D:\Eric\Quad_Data_2017_09_26\\axestest\ch{}.dat".format(i))
    
    Data_Analysis().PlotSingles(single_channels)
    #Data_Analysis().UnpackData("D:\\nEXO\\2017_06_19\\" + "ped.dat")
    #Data_Analysis().Missing_Packet_Check("D:\\Eric\\Packets\\")
    #Data_Analysis().Seperate_Packets("D:\\Eric\\Packets\\", 4, 4)
