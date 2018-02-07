# -*- coding: utf-8 -*-
"""
File Name: read_mean.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 3/9/2016 7:12:33 PM
Last modified: 10/13/2016 10:09:29 AM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl

import openpyxl as px
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from openpyxl import Workbook
import numpy as np
import os
from scripts.int_dac_fit import int_dac_fit
from scripts.fpga_dac_fit import fpga_dac_fit
from scripts.linear_fit_m import linear_fit
import matplotlib.pyplot as plt
from scripts.raw_process2 import make_title
import sys
import glob
import pickle
import warnings
from user_settings import user_editable_settings
settings = user_editable_settings()


####################################################################################################################################################
#Loads the existing pulse data spreadsheet and organizes the data for a single sheet into a 3 dimensional matrix
#Returns a 3D matrix 16 channels wide, however many chips long, and the third dimension is the pulse height for a given dac value
def read_gain(filepath,sheetname,dac_steps):

    W = px.load_workbook(filepath, read_only = True)
    p = W.get_sheet_by_name(name = sheetname)
    
    dacmean=[]
    chip_num = settings.chip_num
    for dac_step in range(dac_steps):
        for chip_id in range(chip_num):
            for chn in range(16):
                vl = p['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dac_step*(chip_num+3)))].value
                dacmean.append(vl)
    dacmean = np.resize(dacmean, [dac_steps, chip_num,16] )
    return dacmean

####################################################################################################################################################
#This function takes an existing plot and some data and does analysis of interest and annotates the existing plot with the results
#It needs the plot, the stats (the list to analyze), a multiplier (to turn the electrons unit into "millions of electrons" for example),
#the subtitle desired, chip number and the written out units to report the slopes in
def plot_stats(plt,stats,multiplier,title,chip_id,unit):

#Where on the plot to put these annotations
    subtitle_text = (.5,.94)
    average_text = (.65,.25)
    maximum_text = (.65,.20)
    minimum_text = (.65,.15)
    std_text = (.65,.10)
    gain_text = (.65,.05)

#Since the list of slopes can contain more than one chip's worth, this isolates the 16 slopes we care about for this chip plot and makes
#it an array for easier analysis below.  It also multiplies the numbers right away so they're in the format we want
    slope_stat = np.multiply(np.array(stats[16*(chip_id)+0:16*(chip_id)+16]), multiplier)


#Depending on whether you're looking at electrons or millivolts, you'll want a certain amount of digits for the number.  This cleans it up
#so you have enough significant figures, but not so much that it clogs up the plot
    if (slope_stat[0] > 1000):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=RuntimeWarning)
            average = str(int(np.nanmean(slope_stat)))
        maximum = str(int(max(slope_stat)))
        minimum = str(int(min(slope_stat)))
        std_dev = str(round(np.std(slope_stat),2))
    else:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=RuntimeWarning)
            average = str(round(np.nanmean(slope_stat),2))
        maximum = str(round(max(slope_stat),2))
        minimum = str(round(min(slope_stat),2))
        std_dev = str(round(np.std(slope_stat),4))

#Self explanatory, it takes the statistics found above and plots them in the designated place.  'axes fraction' lets you give the coordinates
#in terms of the axes, where 0 is one extreme and 1 is the other.  Without that, the coordinates would be in data format, which is harder to use
    plt.annotate(title[0:64]+'\n'+title[64:150], xy=subtitle_text,  xycoords='axes fraction',
    xytext=subtitle_text, textcoords='axes fraction',
    horizontalalignment='center', verticalalignment='center',
    )

    plt.annotate('Average gain is '+average, xy=average_text,  xycoords='axes fraction',
    xytext=average_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Maximum gain is "+maximum, xy=maximum_text,  xycoords='axes fraction',
    xytext=maximum_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Minimum gain is "+minimum, xy=minimum_text,  xycoords='axes fraction',
    xytext=minimum_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Standard Deviation of gains is "+std_dev, xy=std_text,  xycoords='axes fraction',
    xytext=std_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Gain units in ADC counts per "+unit, xy=gain_text,  xycoords='axes fraction',
    xytext=gain_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

####################################################################################################################################################
#Opens an existing pulse data spreadsheet, and for every chip/channel/configuration, analyzes the increase in pulse height
#for increasing DAC values.  It fits this line and plots all the data in the given directories.  It also creates a spreadsheet
#for later RMS/ENC analysis, and populates it with the gain data, so the RMS and mean in ADC counts can be converted to electrons or mV
def get_gain_results(test_dir, pulse_filepath, rms_filepath, intdac, temp):


#Utilize existing knowledge on what each DAC step corresponds to in mV.  The below file must be in the same directory as the .py file
    fpga_vlt_slope = fpga_dac_fit(path = "./scripts/R16_8_4_2_1_77iii_stength.xlsx",ideal_flg = 1) 
    ln2_int_vlt_slope = int_dac_fit(1)
    rt_int_vlt_slope = int_dac_fit(0)

#By looking at the name of the spreadsheet, determine if this test was done by the FPGA or internal DAC.  These distinctions are important
#because 'one DAC step' means something different for the FPGA vs. internal DAC, which is different at Liquid Nitrogen and Room Temperature
    if (intdac == 0):
        vlt_slope = fpga_vlt_slope
        dac_steps = 32
        plot_folder_name = "\\fpga_pulse_fits\\" 
        print ("External FPGA DAC Pulse")
    elif (intdac == 1 and temp == "RT"):
        vlt_slope = ln2_int_vlt_slope
        dac_steps = 64
        plot_folder_name = "\\int_pulse_fits\\" 
        print ("Internal DAC Pulse")
    elif (intdac == 1 and (temp == "LN" or temp == "LXe")):
        vlt_slope = rt_int_vlt_slope
        dac_steps = 64
        plot_folder_name = "\\int_pulse_fits\\" 
        print ("Internal DAC Pulse")
    else:
        sys.exit("Wrong file name")
        
#Get the basics about the test directory we're in

    config_file = (glob.glob(test_dir + "*chip_settings*")[0])
        
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
        baseline = (data["base"])
        gain = (data["gain"])
        gain_num = float(gain[0:4])
        peak = (data["peak"])
        
    sheet_title = gain[0:4] + "," + peak[0:3]
    print (sheet_title)

    
#Create a new workbook for RMS/ENC data
    if (os.path.isfile(rms_filepath)):
        RMS_wb = px.load_workbook(rms_filepath)
        test_titles = [sheet_title+",200", sheet_title+",900"]
        for title in test_titles:
            if (title in RMS_wb.get_sheet_names()):
                RMS_wb.remove_sheet(RMS_wb[title])   
    else:
        RMS_wb = Workbook()

 
#Now that we know which directory, we can create the folder there for plots
    result_path= test_dir + plot_folder_name

    try: 
        os.makedirs(result_path)
    except OSError:
        if os.path.exists(result_path):
            pass


#This function takes in the filepath to the pulse sheet, the sheetname you want analyzed, total number of chips and DAC steps
#(which changes if using FPGA or Interna DAC) and returns the entire dataset for the sheet, organized in a 3D matrix
    chip_num = settings.chip_num
    dacmean = read_gain(pulse_filepath,sheet_title,dac_steps)

    mvslope = []
    eslope = []

    for chip_id in range(chip_num):

#Each chip gets its own plot.  Because you're switching back and forth between two plots, I found this to be the easiest way
        plt.figure(0, figsize=(12,8))
        electron_plot = plt.subplot(1,1,1)

        plt.figure(1, figsize=(12,8))
        mv_plot = plt.subplot(1,1,1)

        electron_plot.figsize=(9,6)
        mv_plot.figsize=(9,6)

        electron_plot.set_xlabel("Test Charge Injection (millions of electrons)")

#So there's not so many zeroes on the electron plt
        electron_plot.ticklabel_format(axis='x', style='sci', scilimits=(-2,2))

        mv_plot.set_xlabel("Estimated Preamplifier Output (mV)")

        
        electron_plot.set_ylabel("ADC counts")
        electron_plot.set_title("Chip %d ADC Output for various injected charges"%(chip_id))

        mv_plot.set_ylabel("ADC counts")
        mv_plot.set_title("Chip %d ADC Output for estimated preamplifier output"%(chip_id))

#Name of the file to save it as
        electron_path = result_path + "electrons_chn_chip"+str(chip_id)
        mv_path = result_path + "mv_chn_chip"+str(chip_id)
        

        for chn in range(16):

#This returns the progressively increasing pulse height for a given channel in ADC counts
#It's a one dimensional list with the same alements as there were DAC steps, starts at DAC step 1
#Because DAC step 0 means no pulse.
            adc_np = dacmean[1:,chip_id,chn]

#This function does all the linear analysis and will return the slope, y-intercept, x-values, and ADC counts (unchanged from when
#passed to the function).  It takes in the chip and channel, which come from the for loop, the adc count which acts as the y-values
#as obtained in the previous line, the slope that turns "DAC number" into the desired x-value.  The 1 means you get back the injected
#charge, in which case the "gain" parameter is meaningless.  The 0 means that the function will return the estimated mV level of the
#preamp output instead of the known charge input.  This estimation is based on the gain that is passed in.
            a,econstant, adc_counts, electrons = linear_fit(chip_id, chn, adc_np,vlt_slope,1,gain_num)
            c,mvconstant, adc_counts, mv = linear_fit(chip_id, chn, adc_np,vlt_slope,0,gain_num)
            
#These are lists of all the gains for each chip and channel
            eslope.append(a)
            mvslope.append(c)

#If estimating the voltage of the preamp output, the baseline has to be taken into account.  This takes the Y-intercept (which by
#definition must be the baseline) and adds it to all the mV readings, since the actual voltage of the preamp at any given time is
#Voltage = Gain(mV/charge) * Input(charge) + Baseline (mV)
            for i in range(len(mv)):
                mv[i] = mv[i]+mvconstant
         
            electron_plot.scatter(electrons, adc_counts, marker='.')
            electron_plot.plot(electrons, adc_counts)
            
            mv_plot.scatter(mv, adc_counts, marker='.')
            mv_plot.plot(mv, adc_counts)

#Once the entire chip has been run through, the final plot processing happens.  This function will fill in all the data analysis of
#interest.  It takes in which plot to analyze, which data to give it, the multiplier (for example the electron data will be multiplied
#by 1 million), the subtitle of the plot, which is the same as the previously found title, the chip number, and the english representation
#of what the units of the slop of the plot are
        title = ("Baseline = {}, Gain = {}, Peaking Time = {}".format(baseline, gain, peak))

        plot_stats(electron_plot,eslope,1e6,title,chip_id,"million electrons")
        
#This is just for the electron plot, it looks weird with a tick that says negative electrons
        xticks = electron_plot.xaxis.get_major_ticks()
        xticks[0].label1.set_visible(False)

#Now that the plot is done, it's selected, saved and cleared, because with this figure(#) method, the next iteration wont create a new plot
        plt.figure(0)
        plt.savefig (electron_path+".jpg")
        plt.clf()


        plot_stats(mv_plot,mvslope,1,title,chip_id,"mV")

        plt.figure(1)
        plt.savefig (mv_path+".jpg")
        plt.clf()

#After all the plots are done, it's time to create the new spreadsheet for ENC analysis  First, the collection of each channel's gain is organized
    eslope = np.resize(eslope, [chip_num,16])
    mvslope = np.resize(mvslope, [chip_num,16])

#Two sheets need to be created for each configuration, because the RMS test goes for 200 mV AND 900 mV baselines, while the one analyzed above only
#did 200 mV.  The gain for each channel is valid at both, so it can simply be copied to both
    RMS_ws1 = RMS_wb.create_sheet()
    RMS_ws2 = RMS_wb.create_sheet()

    RMS_ws1.title = sheet_title+",200"
    RMS_ws2.title = sheet_title+",900"

#This formats the chunk of the spreadsheet to be used for these values.  This function is part of another .py file (right now it's called raw_process2)
#It fills in the title, and labels "Channel 1, Channel 2, etc...) as well as "Chip 1, Chip 2, etc..."  The first number is the "chunk" you want to use,
#The second is the number of chips, the third argument is the title of the chunk, and the fourth is the worksheet
    make_title(3,chip_num,"Gain in ADC counts/mV",RMS_ws1)
    make_title(3,chip_num,"Gain in ADC counts/mV",RMS_ws2)
    make_title(4,chip_num,"Gain in ADC counts/electron",RMS_ws1)
    make_title(4,chip_num,"Gain in ADC counts/electron",RMS_ws2)
   
#For each chip/channel on both 200 and 900 sheets, it just places the gain data.  Format for that sheet is always 
#chip_id+4+(chunk*(chip_num+3)) where the chunk should match up with the ones titled above

    for chip_id in range(chip_num):
        for chn in range(16):
                RMS_ws1['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(3*(chip_num+3)))].value = mvslope[chip_id][chn]
                RMS_ws2['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(3*(chip_num+3)))].value = mvslope[chip_id][chn]

                RMS_ws1['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(4*(chip_num+3)))].value = eslope[chip_id][chn]
                RMS_ws2['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(4*(chip_num+3)))].value = eslope[chip_id][chn]

#Save after every sheet, which is useful when debugging
        RMS_wb.save(filename = rms_filepath)

#After every sheet is done, delete that first empty sheet and save
    if ("Sheet" in RMS_wb.get_sheet_names()):
        RMS_wb.remove_sheet(RMS_wb["Sheet"])   
    RMS_wb.save(filename = rms_filepath)


#Part of the program that runs, change filenames and directories and chip numbers as needed
#root_path = "D:\\femb_3\\2016_10_07"
#pulse_filename = "RF_Pulse_Data.xlsx"
#rms_filename = "R_RMS_Data.xlsx"
#chip_num = 4
#get_gain_results(root_path,pulse_filename,rms_filename,chip_num)


